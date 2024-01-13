import openpyxl

from openpyxl.styles import PatternFill

def row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)

def column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_column)

def read_data(file, sheetname, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rowno,columnno).value

def write_data(file, sheetname, rowno, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rowno,columnno).value = data
    workbook.save(file)