import openpyxl
'''
file =  "J:\Selenium_test_data.xlsx"
workbook = openpyxl.load_workbook(file)

# for multiple sheets
sheet = workbook["Sheet1"]

# for a single sheet in excel
# sheet = workbook.active

for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "WELCOME"                # For writing value into excel
        # data = sheet.cell(r, c).value                     # For reading and saving value into excel
# For saving the file
workbook.save(file)'''

# For writing different data into excel
file =  "J:\Selenium_test_data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = "WELCOME"
sheet.cell(1, 2).value = 2
sheet.cell(1, 3).value = "india"

sheet.cell(2, 1).value = "What"
sheet.cell(2, 2).value = 4
sheet.cell(2, 3).value = "it"

workbook.save(file)