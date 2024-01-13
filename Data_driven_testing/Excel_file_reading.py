import openpyxl

''' File ---> Workbook ---> Sheet --->  Rows ----> Cells'''
file =  "J:\Selenium_test_data.xlsx"
workbook = openpyxl.load_workbook(file)    # loading workbook from file
sheet = workbook["Sheet1"]

# To count rows in excel
rows = sheet.max_row
# To count columns in excel
cols = sheet.max_column

# Reading all the rows and columns
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end=" | ")
    print()